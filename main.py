from __future__ import annotations

import csv
import shutil
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse
from starlette.background import BackgroundTask
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas


app = FastAPI(title="Charger Report Converter API")


# =========================
# Data model
# =========================

@dataclass
class ChargeCycle:
    cycle_number: str
    date_time: str
    charge_time: str
    end_code: str
    end_amps: str
    end_volts: str
    ah: str


@dataclass
class ReportData:
    charger_info: Dict[str, str]
    settings_1: Dict[str, str]
    settings_2: Dict[str, str]
    algorithm_1: Dict[str, str]
    algorithm_2: Dict[str, str]
    f_disables: Dict[str, str]
    phases: List[Dict[str, str]]
    abnormal_cycles: List[ChargeCycle]
    hot_disconnects: List[ChargeCycle]
    equalize_cycles: List[ChargeCycle]
    all_cycles: List[ChargeCycle]


# =========================
# Input readers
# =========================

def read_csv_rows(path: Path) -> list[list[str]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return [[str(cell).strip() for cell in row] for row in csv.reader(f)]


def read_excel_rows(path: Path) -> list[list[str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell).strip() for cell in row])
    return rows


def read_rows(path: Path) -> list[list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_excel_rows(path)
    raise ValueError(f"Unsupported file type: {suffix}")


# =========================
# Parsing helpers
# =========================

def normalize(text: str) -> str:
    return " ".join(str(text).strip().split()).lower()


def row_has_text(row: list[str]) -> bool:
    return any(str(cell).strip() for cell in row)


def first_nonempty(row: list[str]) -> str:
    for cell in row:
        if str(cell).strip():
            return str(cell).strip()
    return ""


def find_row(rows: list[list[str]], first_cell: str) -> int:
    target = normalize(first_cell)
    for i, row in enumerate(rows):
        if normalize(first_nonempty(row)) == target:
            return i
    raise ValueError(f"Could not find section: {first_cell}")


def pair_dict(headers: list[str], values: list[str]) -> Dict[str, str]:
    result: Dict[str, str] = {}
    n = max(len(headers), len(values))
    for i in range(n):
        h = headers[i].strip() if i < len(headers) else ""
        v = values[i].strip() if i < len(values) else ""
        if h:
            result[h] = v
    return result


def find_header_value_block(
    rows: list[list[str]],
    start_idx: int,
    header_first_cell: str,
) -> Dict[str, str]:
    target = normalize(header_first_cell)
    for i in range(start_idx, len(rows) - 1):
        if normalize(first_nonempty(rows[i])) == target:
            return pair_dict(rows[i], rows[i + 1])
    return {}


def parse_phase_block(rows: list[list[str]], start_idx: int, phase_name: str) -> Dict[str, str]:
    target = normalize(f"{phase_name} Control")
    for i in range(start_idx, len(rows) - 1):
        if normalize(first_nonempty(rows[i])) == target:
            data = pair_dict(rows[i], rows[i + 1])
            data["_phase"] = phase_name
            return data
    return {"_phase": phase_name}


# =========================
# Parse report
# =========================

def parse_report(rows: list[list[str]]) -> ReportData:
    charger_idx = find_row(rows, "Charger Info")
    charger_info = pair_dict(rows[charger_idx + 1], rows[charger_idx + 2])

    settings1_idx = find_row(rows, "Settings 1")
    settings_1 = pair_dict(rows[settings1_idx + 1], rows[settings1_idx + 2])

    settings2_idx = find_row(rows, "Settings 2")
    settings_2 = pair_dict(rows[settings2_idx + 1], rows[settings2_idx + 2])

    algo_idx = find_row(rows, "Algorithm")
    algorithm_1 = find_header_value_block(rows, algo_idx + 1, "Trip Point Voltage")
    algorithm_2 = find_header_value_block(rows, algo_idx + 1, "Equalize Type")
    f_disables = find_header_value_block(rows, algo_idx + 1, "F2 Disable")

    phases = [
        parse_phase_block(rows, algo_idx + 1, "Phase 1"),
        parse_phase_block(rows, algo_idx + 1, "Phase 2"),
        parse_phase_block(rows, algo_idx + 1, "Phase 3"),
        parse_phase_block(rows, algo_idx + 1, "Phase 4"),
    ]

    cycles_idx = find_row(rows, "Cycles")
    headers = rows[cycles_idx + 1]
    header_map = {normalize(h): idx for idx, h in enumerate(headers) if str(h).strip()}

    def get_col(row: list[str], name: str) -> str:
        idx = header_map.get(normalize(name))
        if idx is None or idx >= len(row):
            return ""
        return row[idx].strip()

    abnormal_cycles: list[ChargeCycle] = []
    hot_disconnects: list[ChargeCycle] = []
    equalize_cycles: list[ChargeCycle] = []
    all_cycles: list[ChargeCycle] = []

    for r in rows[cycles_idx + 2:]:
        if not row_has_text(r):
            continue

        cycle_number = get_col(r, "Cycle Number")
        if not cycle_number.isdigit():
            continue

        cycle = ChargeCycle(
            cycle_number=cycle_number,
            date_time=get_col(r, "Date/Time"),
            charge_time=get_col(r, "Charge Time"),
            end_code=get_col(r, "End Code"),
            end_amps=get_col(r, "End Amps"),
            end_volts=get_col(r, "End Volts"),
            ah=get_col(r, "AH"),
        )

        all_cycles.append(cycle)

        end_code_upper = cycle.end_code.upper()

        if cycle.end_code.startswith("F"):
            abnormal_cycles.append(cycle)

        if "BATTERY DISCONNECT" in end_code_upper:
            hot_disconnects.append(cycle)

        if "EQ CHARGE COMPLETE" in end_code_upper:
            equalize_cycles.append(cycle)

    return ReportData(
        charger_info=charger_info,
        settings_1=settings_1,
        settings_2=settings_2,
        algorithm_1=algorithm_1,
        algorithm_2=algorithm_2,
        f_disables=f_disables,
        phases=phases,
        abnormal_cycles=abnormal_cycles,
        hot_disconnects=hot_disconnects,
        equalize_cycles=equalize_cycles,
        all_cycles=all_cycles,
    )


# =========================
# PDF writer
# =========================

class PDFReport:
    def __init__(self, out_path: Path):
        self.c = canvas.Canvas(str(out_path), pagesize=letter)
        self.width, self.height = letter
        self.left = 0.6 * inch
        self.top = self.height - 0.6 * inch
        self.y = self.top
        self.bottom = 0.6 * inch

    def ensure_space(self, needed_inches: float) -> None:
        if self.y - needed_inches * inch < self.bottom:
            self.new_page()

    def new_page(self) -> None:
        self.c.showPage()
        self.y = self.top

    def title(self, text: str) -> None:
        self.ensure_space(0.40)
        self.c.setFont("Helvetica", 13)
        self.c.drawString(self.left, self.y, text)
        self.y -= 0.24 * inch

    def draw_label_value(self, label: str, value: str) -> None:
        self.ensure_space(0.26)
        label_text = f"{label}: "

        self.c.setFont("Helvetica", 13)
        self.c.drawString(self.left, self.y, label_text)
        label_width = self.c.stringWidth(label_text, "Helvetica", 13)

        self.c.setFont("Helvetica", 13)
        self.c.drawString(self.left + label_width, self.y, value or "")
        self.y -= 0.22 * inch

    def gap(self, size: float = 0.10) -> None:
        self.y -= size * inch

    def wrap_text_to_width(
        self,
        text: str,
        max_width: float,
        font_name: str,
        font_size: int,
    ) -> list[str]:
        text = str(text or "")
        if not text:
            return [""]

        paragraphs = text.split("\n")
        wrapped_lines: list[str] = []

        for paragraph in paragraphs:
            words = paragraph.split()
            if not words:
                wrapped_lines.append("")
                continue

            current = words[0]
            for word in words[1:]:
                candidate = f"{current} {word}"
                if self.c.stringWidth(candidate, font_name, font_size) <= max_width:
                    current = candidate
                else:
                    wrapped_lines.append(current)
                    current = word
            wrapped_lines.append(current)

        return wrapped_lines

    def draw_boxed_table(
        self,
        headers: list[str],
        rows: list[list[str]],
        widths: list[float],
        header_font_size: int = 13,
        body_font_size: int = 11,
        cell_padding: float = 0.04 * inch,
        min_row_height: float = 0.35 * inch,
        line_spacing: float = 0.16 * inch,
    ) -> None:
        usable_widths = [max(w - (2 * cell_padding), 0.1 * inch) for w in widths]
        total_width = sum(widths)
        header_height = max(min_row_height, 0.40 * inch)

        wrapped_rows: list[list[list[str]]] = []
        row_heights: list[float] = []

        for row in rows:
            wrapped_row: list[list[str]] = []
            max_lines = 1

            for value, usable_width in zip(row, usable_widths):
                lines = self.wrap_text_to_width(value, usable_width, "Helvetica", body_font_size)
                wrapped_row.append(lines)
                max_lines = max(max_lines, len(lines))

            row_height = max(min_row_height, ((max_lines - 1) * line_spacing) + 0.30 * inch)
            wrapped_rows.append(wrapped_row)
            row_heights.append(row_height)

        def draw_table_header(y_top: float) -> float:
            x = self.left
            self.c.setFont("Helvetica-Bold", header_font_size)

            for text, w in zip(headers, widths):
                header_lines = self.wrap_text_to_width(
                    text,
                    w - (2 * cell_padding),
                    "Helvetica-Bold",
                    header_font_size,
                )
                header_block_height = (len(header_lines) - 1) * line_spacing
                first_line_y = y_top - ((header_height - header_block_height) / 2.0) - (header_font_size * 0.30)

                for line_index, line in enumerate(header_lines):
                    y_text = first_line_y - (line_index * line_spacing)
                    self.c.drawString(x + cell_padding, y_text, line)

                x += w

            return y_top - header_height

        row_index = 0
        first_page = True

        while row_index < len(wrapped_rows):
            if not first_page:
                self.new_page()
            first_page = False

            self.ensure_space((header_height / inch) + 0.5)

            y_top = self.y
            available_height = y_top - self.bottom

            used_height = header_height
            start_index = row_index
            while row_index < len(row_heights):
                if used_height + row_heights[row_index] > available_height:
                    break
                used_height += row_heights[row_index]
                row_index += 1

            if start_index == row_index:
                self.new_page()
                y_top = self.y
                available_height = y_top - self.bottom
                used_height = header_height
                while row_index < len(row_heights):
                    if used_height + row_heights[row_index] > available_height:
                        break
                    used_height += row_heights[row_index]
                    row_index += 1

            end_index = row_index
            page_row_heights = row_heights[start_index:end_index]
            page_rows = wrapped_rows[start_index:end_index]
            total_height = header_height + sum(page_row_heights)
            y_bottom = y_top - total_height

            self.c.rect(self.left, y_bottom, total_width, total_height)

            x = self.left
            for w in widths[:-1]:
                x += w
                self.c.line(x, y_bottom, x, y_top)

            y_cursor = y_top - header_height
            self.c.line(self.left, y_cursor, self.left + total_width, y_cursor)

            for rh in page_row_heights[:-1]:
                y_cursor -= rh
                self.c.line(self.left, y_cursor, self.left + total_width, y_cursor)

            current_top = draw_table_header(y_top)

            self.c.setFont("Helvetica", body_font_size)
            for wrapped_row, row_height in zip(page_rows, page_row_heights):
                x = self.left
                for cell_lines, w in zip(wrapped_row, widths):
                    line_count = max(len(cell_lines), 1)
                    block_height = (line_count - 1) * line_spacing
                    first_line_y = current_top - ((row_height - block_height) / 2.0) - (body_font_size * 0.30)

                    for line_index, line in enumerate(cell_lines):
                        y_text = first_line_y - (line_index * line_spacing)
                        self.c.drawString(x + cell_padding, y_text, line)

                    x += w

                current_top -= row_height

            self.y = y_bottom - 0.15 * inch

    def save(self) -> None:
        self.c.save()


def render_cycle_table_page(pdf: PDFReport, title: str, cycles: list[ChargeCycle]) -> None:
    pdf.new_page()
    pdf.title(f"{title}: {len(cycles)}")

    if cycles:
        headers = ["Cycle Number", "Date/Time", "Charge Time", "End Code", "End Amps", "End Volts", "AH"]
        widths = [
            1.05 * inch,
            1.35 * inch,
            1.10 * inch,
            1.45 * inch,
            1.05 * inch,
            1.05 * inch,
            0.55 * inch,
        ]

        rows: list[list[str]] = []
        for c in cycles:
            rows.append([
                c.cycle_number,
                c.date_time,
                c.charge_time,
                c.end_code,
                c.end_amps,
                c.end_volts,
                c.ah,
            ])

        pdf.draw_boxed_table(
            headers=headers,
            rows=rows,
            widths=widths,
            header_font_size=9,
            body_font_size=9,
            min_row_height=0.42 * inch,
            line_spacing=0.16 * inch,
        )
    else:
        pdf.draw_label_value("Result", "None")


def render_report(data: ReportData, output_pdf: Path) -> None:
    pdf = PDFReport(output_pdf)

    pdf.title("Charger Info")
    for key in [
        "Model Number",
        "Serial Number",
        "Customer ID",
        "Factory ID",
        "Voltage Rating",
        "AH Rating",
        "Type",
        "Area",
        "Location",
        "S/W Version",
    ]:
        pdf.draw_label_value(key, data.charger_info.get(key, ""))
    pdf.gap()

    pdf.title("Settings")
    pdf.draw_label_value("Algorithm Number", data.settings_1.get("Algorithm", ""))
    for key in [
        "Shunt Rating",
        "Max Power",
        "Set Amps",
        "Set Volts",
        "Set Time",
        "Battery Com Mode",
        "Cable Resistance",
        "CANbus baud rate",
    ]:
        pdf.draw_label_value(key, data.settings_1.get(key, ""))

    for key in [
        "AH Rating 2",
        "AH Rating 3",
        "Auto Start Mode",
        "TOD Start",
        "Delayed Start",
        "Charge Factor",
        "Cool Down Direction",
        "Cool Down Time",
        "Watering Mode",
        "Watering Cycles",
        "Auto Refresh Time",
        "Peak Start",
        "Peak End",
    ]:
        pdf.draw_label_value(key, data.settings_2.get(key, ""))

    pdf.new_page()
    pdf.title("Algorithm Settings")

    for key in [
        "Trip Point Voltage",
        "Cutoff Voltage",
        "OK To Charge Temp",
        "Low Charge Temp",
        "No Charge Temp",
        "Special Code",
    ]:
        pdf.draw_label_value(key, data.algorithm_1.get(key, ""))

    pdf.gap(0.14)

    for key in [
        "Equalize Type",
        "Equalize Current",
        "Equalize By Cycles",
        "Equalize By Day",
        "Equalize Delay",
        "Equalize Time",
    ]:
        pdf.draw_label_value(key, data.algorithm_2.get(key, ""))

    pdf.gap(0.14)

    for key in ["F2 Disable", "F3 Disable", "F4 Disable", "F6 Disable"]:
        pdf.draw_label_value(key, data.f_disables.get(key, ""))

    pdf.gap()

    criteria_headers = ["Phase", "Control", "End Criteria", "Amps", "Volts", "Time", "Time Code"]
    criteria_widths = [
        0.95 * inch,
        1.00 * inch,
        1.25 * inch,
        0.65 * inch,
        0.65 * inch,
        0.75 * inch,
        1.00 * inch,
    ]

    criteria_rows: list[list[str]] = []
    for p in data.phases:
        phase = p["_phase"]
        criteria_rows.append([
            phase,
            p.get(f"{phase} Control", ""),
            p.get(f"{phase} End", ""),
            p.get(f"{phase} Amps", ""),
            p.get(f"{phase} V/C", ""),
            p.get(f"{phase} Time", ""),
            p.get(f"{phase} Time Code", ""),
        ])

    pdf.draw_boxed_table(
        headers=criteria_headers,
        rows=criteria_rows,
        widths=criteria_widths,
        header_font_size=9,
        body_font_size=9,
        min_row_height=0.35 * inch,
        line_spacing=0.16 * inch,
    )

    pdf.gap(0.25)

    for label, meaning in [
        ("GV", "Greater than Volts"),
        ("LA", "Less than Amps"),
        ("DV", "dVdt/dIdT"),
        ("T", "Time"),
        ("LV", "Less than Volts"),
    ]:
        pdf.draw_label_value(label, meaning)

    render_cycle_table_page(pdf, "Abnormal Charge Cycles", data.abnormal_cycles)
    render_cycle_table_page(pdf, "Hot Disconnects", data.hot_disconnects)
    render_cycle_table_page(pdf, "Equalize Cycles", data.equalize_cycles)
    render_cycle_table_page(pdf, "Charger History", data.all_cycles)

    pdf.save()


# =========================
# Convert function
# =========================

def convert_file_to_pdf(input_path: str | Path, output_dir: str | Path | None = None) -> Path:
    input_path = Path(input_path)
    rows = read_rows(input_path)
    report = parse_report(rows)

    if output_dir is None:
        output_dir = input_path.parent

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    output_pdf = output_dir / f"{input_path.stem}_report.pdf"
    render_report(report, output_pdf)
    return output_pdf


# =========================
# FastAPI endpoints
# =========================

@app.get("/")
def health_check() -> dict[str, str]:
    return {"status": "ok", "message": "Charger Report Converter API is running"}


@app.post("/convert-report")
async def convert_report_endpoint(file: UploadFile = File(...)) -> FileResponse:
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    suffix = Path(file.filename).suffix.lower()
    if suffix not in {".csv", ".xlsx", ".xlsm"}:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Please upload a .csv, .xlsx, or .xlsm file.",
        )

    temp_dir_obj = tempfile.TemporaryDirectory()
    temp_dir = Path(temp_dir_obj.name)

    try:
        input_path = temp_dir / file.filename

        with input_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        output_pdf = convert_file_to_pdf(input_path, output_dir=temp_dir)

        return FileResponse(
            path=str(output_pdf),
            media_type="application/pdf",
            filename=output_pdf.name,
            background=BackgroundTask(temp_dir_obj.cleanup),
        )

    except ValueError as exc:
        temp_dir_obj.cleanup()
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        temp_dir_obj.cleanup()
        raise HTTPException(status_code=500, detail=f"Conversion failed: {exc}") from exc


@app.get("/health")
def health() -> JSONResponse:
    return JSONResponse(content={"status": "healthy"})
